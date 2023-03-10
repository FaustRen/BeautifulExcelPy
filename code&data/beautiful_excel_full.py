#%%
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# read csv data
path="./Hotel Reservations.csv"
font_setting="Arial"
save_name="test_excel_haha.xlsx"
# path="./csv檔案路徑"
df = pd.read_csv(path)

def beautifulWorkbook(data_input,output_filename,font_type='Arial'):
    """儲存至xlsx用, 儲存並修改字體
    data_input:         df 變數名稱
    output_filename:    儲存檔案的名稱
    font_type:          儲存後字體
    """
    
    # 建立一個ExcelWriter物件，並設定寫入的檔案路徑和檔名
    # writer = pd.ExcelWriter('output45678.xlsx')
    writer = pd.ExcelWriter(output_filename)
    df=data_input.copy()

    # 將DataFrame寫入Excel檔案中，並設定工作表名稱為「Data」
    df.to_excel(writer, sheet_name='Data', index=False)

    # 取得寫入的工作表物件
    worksheet = writer.sheets['Data']

    # 調整每一格的行列距離，以顯示該格資料
    for idx, col in enumerate(df):
        series = df[col]
        max_len = max((
            series.astype(str).map(len).max(), 
            len(str(series.name))
            ))
        worksheet.set_column(idx, idx, max_len + 2)

    # 設定字體為Arial
    cell_format = writer.book.add_format({'font_name': font_type})
    worksheet.set_column(0, len(df.columns)-1, None, cell_format)

    # 儲存Excel檔案
    writer.save()
# %%
beautifulWorkbook(df, save_name,font_setting)
# %%
